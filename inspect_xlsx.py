import openpyxl
path = "/home/coder/dev/x1/kyle/route_commerce-main/Tuxedo_Corn_2026_Tour_Schedule-3.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"=== SHEET: {name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is not None:
                v = str(cell.value)
                if len(v) > 200:
                    v = v[:200] + "..."
                print(f"  {cell.coordinate}: {v!r}")
    print()
