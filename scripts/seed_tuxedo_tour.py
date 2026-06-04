#!/usr/bin/env python3
"""
seed_tuxedo_tour.py

Parses Tuxedo_Corn_2026_Tour_Schedule-3.xlsx and seeds the `stops` table for
the Tuxedo brand via the admin_create_stops_batch RPC.

Skips:
  - Title / subtitle / legend rows (rows 1-3)
  - Week header rows (col A = "Wk N", col D-J empty)
  - Cross-Dock / Monday OFF rows (col D contains "OFF" or "Cross-Dock")

Joins with the Stop Directory sheet to enrich each stop with:
  - address, phone, contact

Usage:
  python3 scripts/seed_tuxedo_tour.py --dry-run   # show what would be inserted
  python3 scripts/seed_tuxedo_tour.py            # actually insert

Requires:
  - supabase CLI linked to project wnzkhezyhnfzhkhiflrp
"""
import argparse
import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

TUXEDO_BRAND_ID = "64294306-5f42-463d-a5e8-2ad6c81a96de"
YEAR = 2026

MONTH_MAP = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}

DEFAULT_XLSX = (
    "/home/coder/dev/x1/kyle/route_commerce-main/"
    "Tuxedo_Corn_2026_Tour_Schedule-3.xlsx"
)


def parse_excel_date(s):
    """'Jul 22' -> '2026-07-22'"""
    if not s:
        return None
    m = re.match(r"^([A-Za-z]{3})\s+(\d{1,2})$", str(s).strip())
    if not m:
        return None
    mm = MONTH_MAP.get(m.group(1))
    if not mm:
        return None
    return f"{YEAR}-{mm}-{int(m.group(2)):02d}"


def parse_time_range(s):
    """'10:00 AM - 1:00 PM' -> '10:00 AM' (start time)"""
    if not s:
        return ""
    cleaned = re.sub(r"[–—]", "-", str(s)).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    m = re.match(r"^(\d{1,2}:\d{2}\s*[AP]M)", cleaned, re.IGNORECASE)
    return m.group(1).upper().replace(" ", " ") if m else cleaned


def split_city_state(s):
    """'Cheyenne, WY' -> ('Cheyenne', 'WY')"""
    if not s:
        return "", ""
    parts = [p.strip() for p in str(s).split(",")]
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def slugify(s):
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def is_week_header(row):
    a = str(row[0] or "").strip()
    d = str(row[3] or "").strip()
    return re.match(r"^Wk\s", a) and d == ""


def is_off_row(row):
    d = str(row[3] or "").strip()
    return "OFF" in d or "Cross-Dock" in d or "Cross‑Dock" in d


def is_data_row(row):
    d = str(row[3] or "").strip()
    e = str(row[4] or "").strip()
    if not d or not e:
        return False
    if "," not in e:
        return False
    return True


def load(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    schedule = wb["Full Schedule"]
    directory = wb["Stop Directory"]

    # Build Stop Directory lookup: (truck, host_normalized) -> {address, phone, contact, ...}
    dir_map = {}
    for row in directory.iter_rows(min_row=2, values_only=True):
        truck = str(row[0] or "").strip()
        city = str(row[1] or "").strip()
        state = str(row[2] or "").strip()
        host = str(row[3] or "").strip()
        address = str(row[4] or "").strip()
        phone = str(row[5] or "").strip()
        contact = str(row[6] or "").strip()
        if not truck or not host:
            continue
        key = f"{truck}|{host.lower()}"
        dir_map[key] = {
            "city": city, "state": state, "host": host,
            "address": address, "phone": phone, "contact": contact,
        }

    # Read Full Schedule (skip first 3 title/subtitle/legend rows)
    stops = []
    skipped = {"weekHeader": 0, "off": 0, "invalid": 0}
    for row in schedule.iter_rows(min_row=4, values_only=True):
        # Trim to 10 cols
        cells = [("" if v is None else str(v).strip()) for v in row[:10]]
        if is_week_header(cells):
            skipped["weekHeader"] += 1
            continue
        if is_off_row(cells):
            skipped["off"] += 1
            continue
        if not is_data_row(cells):
            skipped["invalid"] += 1
            continue

        wk, region, date_text, day, city_state, host, time, truck, status, notes = cells
        date_iso = parse_excel_date(date_text)
        if not date_iso:
            skipped["invalid"] += 1
            continue
        city, state = split_city_state(city_state)
        if not city:
            skipped["invalid"] += 1
            continue

        # Enrich from directory
        dir_key = f"{truck}|{host.lower()}"
        d = dir_map.get(dir_key)

        stops.append({
            "week": wk,
            "region": region,
            "date": date_iso,
            "day": day,
            "city": city,
            "state": state or (d["state"] if d else ""),
            "location": host,
            "time": parse_time_range(time),
            "time_range": time,
            "truck": truck,
            "status_text": status,
            "notes": notes,
            "address": d["address"] if d and d["address"] else None,
            "phone": d["phone"] if d and d["phone"] else None,
            "contact": d["contact"] if d and d["contact"] else None,
        })

    return stops, skipped, len(dir_map)


def assign_slugs(stops, dry_run):
    used = set()
    if not dry_run:
        out = subprocess.run(
            ["supabase", "db", "query", "--linked",
             f"SELECT slug FROM stops WHERE brand_id = '{TUXEDO_BRAND_ID}';"],
            capture_output=True, text=True, timeout=120,
        )
        # Parse the table output - slugs are in second column between │
        for m in re.finditer(r"│\s*([a-z0-9][a-z0-9-]*)\s*│", out.stdout):
            used.add(m.group(1))

    for s in stops:
        base = f"{slugify(s['city'])}-{s['date']}"
        slug = base
        n = 0
        while slug in used:
            n += 1
            slug = f"{base}-{n}"
        used.add(slug)
        s["slug"] = slug


def to_rpc_row(s):
    return {
        "city": s["city"],
        "state": s["state"],
        "location": s["location"],
        "date": f"{s['date']} 00:00:00+00",
        "time": s["time"],
        "address": s["address"],
        "zip": None,
        "cutoff_time": None,
        # active=true so the stops appear on the public storefront immediately.
        # Matches the behavior of publishStop in src/actions/stops.ts.
        "active": True,
    }


def build_payload_json(batch):
    """Build a clean JSON string for use in a SQL file."""
    return json.dumps(batch, ensure_ascii=False)


def insert_batch(batch):
    """Write SQL to a temp file and execute via --file to avoid shell escaping."""
    payload_json = build_payload_json(batch)
    sql = (
        f"SELECT admin_create_stops_batch("
        f"'{TUXEDO_BRAND_ID}'::uuid, "
        f"$${payload_json}$$::jsonb);\n"
    )
    # Write to temp file
    tmp_path = Path("/tmp/seed_tuxedo_tour.sql")
    tmp_path.write_text(sql, encoding="utf-8")
    try:
        proc = subprocess.run(
            ["supabase", "db", "query", "--linked", "--file", str(tmp_path)],
            capture_output=True, text=True, timeout=300,
        )
    finally:
        tmp_path.unlink(missing_ok=True)
    if proc.returncode != 0:
        raise RuntimeError(f"RPC failed: {proc.stderr[:800]}")
    return proc.stdout


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    args = ap.parse_args()

    if not Path(args.xlsx).exists():
        sys.exit(f"XLSX not found: {args.xlsx}")

    stops, skipped, dir_count = load(args.xlsx)
    assign_slugs(stops, dry_run=args.dry_run)

    print(f"\nParsed {len(stops)} stops "
          f"(skipped: {skipped['weekHeader']} week-headers, "
          f"{skipped['off']} OFF days, {skipped['invalid']} invalid)")
    print(f"Stop Directory: {dir_count} entries loaded for enrichment\n")

    if not stops:
        sys.exit("No stops to insert.")

    print("Sample (first 3):")
    for s in stops[:3]:
        print(f"  {s['date']} {s['time']:10s} | {s['city']:18s}, {s['state']:2s} | "
              f"{s['location'][:35]:35s} | {s['truck']} | {s['status_text']} | {s['slug']}")
        if s["notes"]:
            print(f"    notes:  {s['notes'][:120]}")
        if s["address"]:
            print(f"    addr:   {s['address']}  ph: {s['phone']}  ctc: {s['contact']}")
    print()

    # Show counts by week and region
    by_week = {}
    by_region = {}
    by_truck = {}
    for s in stops:
        by_week[s["week"]] = by_week.get(s["week"], 0) + 1
        by_region[s["region"]] = by_region.get(s["region"], 0) + 1
        by_truck[s["truck"]] = by_truck.get(s["truck"], 0) + 1
    print("By week:", dict(sorted(by_week.items())))
    print("By region:", by_region)
    print("By truck:", by_truck)
    print()

    # Date range
    dates = sorted(s["date"] for s in stops)
    print(f"Date range: {dates[0]} to {dates[-1]}\n")

    if args.dry_run:
        batches = (len(stops) + 49) // 50
        print(f"[DRY RUN] Would insert {len(stops)} stops in {batches} batch(es) of 50.")
        return

    BATCH = 50
    total = 0
    batches = (len(stops) + BATCH - 1) // BATCH
    for i in range(0, len(stops), BATCH):
        batch = [to_rpc_row(s) for s in stops[i:i + BATCH]]
        bnum = i // BATCH + 1
        sys.stdout.write(f"  Inserting batch {bnum}/{batches} ({len(batch)} stops)... ")
        sys.stdout.flush()
        try:
            insert_batch(batch)
            total += len(batch)
            print("OK")
        except Exception as e:
            print("FAIL")
            print(f"    {e}")

    # The batch RPC hardcodes status='draft' on insert. The Tuxedo storefront
    # page only filters on active=true (not status), so active=true is enough
    # to make stops visible. But for consistency with the publishStop server
    # action — which sets both — flip status to 'active' for the rows we just
    # inserted. Slug-based so we only touch stops from this run, not the
    # pre-existing "Olathe" test stop.
    if total > 0:
        slugs = [s["slug"] for s in stops]
        # Build a safe IN list (slug is a text column)
        slug_list = ", ".join(f"'{slug.replace(chr(39), chr(39)+chr(39))}'" for slug in slugs)
        publish_sql = (
            f"UPDATE stops SET status = 'active' "
            f"WHERE brand_id = '{TUXEDO_BRAND_ID}' "
            f"AND slug IN ({slug_list});"
        )
        tmp = Path("/tmp/seed_tuxedo_publish.sql")
        tmp.write_text(publish_sql, encoding="utf-8")
        try:
            subprocess.run(
                ["supabase", "db", "query", "--linked", "--file", str(tmp)],
                capture_output=True, text=True, timeout=120,
            )
            print(f"\n  Published {total} stops (status -> 'active').")
        finally:
            tmp.unlink(missing_ok=True)

    print(f"\nDone. Inserted {total}/{len(stops)} stops for Tuxedo brand.")


if __name__ == "__main__":
    main()
